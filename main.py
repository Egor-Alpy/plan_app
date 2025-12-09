import sys
from datetime import datetime, timedelta
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QPushButton, QLabel, QComboBox,
                             QTableWidget, QTableWidgetItem, QMessageBox,
                             QFileDialog, QHeaderView, QFrame, QScrollArea)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont


class EducationalScheduleApp:
    def __init__(self):
        # Русские названия месяцев
        self.month_names_ru = {
            1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель',
            5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
            9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
        }

        # Праздничные дни России по годам
        self.holidays = {
            2025: ['2025-01-01', '2025-01-02', '2025-01-03', '2025-01-04', '2025-01-06',
                   '2025-01-07', '2025-01-08', '2025-02-23', '2025-03-08', '2025-05-01',
                   '2025-05-02', '2025-05-08', '2025-05-09', '2025-06-12', '2025-06-13',
                   '2025-11-03', '2025-11-04'],
            2026: ['2026-01-01', '2026-01-02', '2026-01-05', '2026-01-06', '2026-01-07',
                   '2026-01-08', '2026-01-09', '2026-02-23', '2026-03-09', '2026-05-01',
                   '2026-05-09', '2026-05-11', '2026-06-12', '2026-11-04'],
            2027: ['2027-01-01', '2027-01-04', '2027-01-05', '2027-01-06', '2027-01-07',
                   '2027-01-08', '2027-02-22', '2027-02-23', '2027-03-08', '2027-05-03',
                   '2027-05-10', '2027-06-14', '2027-11-04'],
            2028: ['2028-01-03', '2028-01-04', '2028-01-05', '2028-01-06', '2028-01-07',
                   '2028-02-23', '2028-03-08', '2028-05-01', '2028-05-09', '2028-06-12',
                   '2028-11-04']
        }

    def get_monday_of_week(self, date):
        days_since_monday = date.weekday()
        return date - timedelta(days=days_since_monday)

    def is_holiday(self, date):
        year = date.year
        date_str = date.strftime('%Y-%m-%d')
        return date_str in self.holidays.get(year, [])

    def is_working_day(self, date):
        return date.weekday() < 5 and not self.is_holiday(date)

    def calculate_academic_weeks(self, start_date, weeks_float):
        current_date = start_date
        working_days_needed = int(weeks_float * 5)
        working_days_count = 0
        schedule_days = []

        while working_days_count < working_days_needed:
            if self.is_working_day(current_date):
                schedule_days.append(current_date)
                working_days_count += 1
            current_date += timedelta(days=1)

        while not self.is_working_day(current_date):
            current_date += timedelta(days=1)

        return schedule_days, current_date

    def generate_schedule(self, periods_data, start_year):
        start_date = datetime(start_year, 9, 1)
        current_date = self.get_monday_of_week(start_date)

        generated_schedule = []

        for row in periods_data:
            year = int(row['Год'])
            semester = int(row['Семестр'])
            activity_type = row['Тип']
            weeks = float(row['Недели'])

            period_days, next_date = self.calculate_academic_weeks(current_date, weeks)

            period_info = {
                'year': year,
                'semester': semester,
                'type': activity_type,
                'weeks': weeks,
                'start_date': current_date,
                'end_date': period_days[-1] if period_days else current_date,
                'days': period_days
            }

            generated_schedule.append(period_info)
            current_date = next_date

        return generated_schedule

    def create_excel_file(self, generated_schedule, start_year, program_type):
        wb = Workbook()

        program_years = 2 if "Ординатура" in program_type else 3

        # ===== СТИЛИ =====
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        title_font = Font(name='Calibri', size=16, bold=True, color='1976D2')
        year_title_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        legend_header_font = Font(name='Calibri', size=12, bold=True, color='1976D2')
        legend_font = Font(name='Calibri', size=10, color='000000')
        data_font = Font(name='Calibri', size=10, color='000000')

        thin_border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )

        thick_border = Border(
            left=Side(style='medium', color='90CAF9'),
            right=Side(style='medium', color='90CAF9'),
            top=Side(style='medium', color='90CAF9'),
            bottom=Side(style='medium', color='90CAF9')
        )

        activity_fills = {
            'Т': PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid"),
            'Э': PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid"),
            'П': PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
            'У': PatternFill(start_color="B2EBF2", end_color="B2EBF2", fill_type="solid"),
            'ПА': PatternFill(start_color="FFCC80", end_color="FFCC80", fill_type="solid"),
            'ГИА': PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid"),
            'Г': PatternFill(start_color="F8BBD0", end_color="F8BBD0", fill_type="solid"),
            'Д': PatternFill(start_color="D1C4E9", end_color="D1C4E9", fill_type="solid"),
            'К': PatternFill(start_color="FFE082", end_color="FFE082", fill_type="solid"),
        }

        weekend_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        holiday_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        header_fill = PatternFill(start_color="64B5F6", end_color="64B5F6", fill_type="solid")
        month_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        month_fill_alt = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        year_header_fill = PatternFill(start_color="42A5F5", end_color="42A5F5", fill_type="solid")
        legend_header_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        # ===== ЛИСТ 1: УСЛОВНЫЕ ОБОЗНАЧЕНИЯ =====
        ws_legend = wb.active
        ws_legend.title = "Условные обозначения"

        current_row = 1

        ws_legend.merge_cells(f'A{current_row}:F{current_row}')
        ws_legend[f'A{current_row}'] = f"КАЛЕНДАРНЫЙ УЧЕБНЫЙ ГРАФИК {start_year}-{start_year + program_years} г."
        ws_legend[f'A{current_row}'].font = title_font
        ws_legend[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws_legend[f'A{current_row}'].fill = legend_header_fill
        ws_legend[f'A{current_row}'].border = thick_border
        ws_legend.row_dimensions[current_row].height = 35
        current_row += 1

        current_row += 1

        ws_legend.merge_cells(f'A{current_row}:F{current_row}')
        ws_legend[f'A{current_row}'] = "УСЛОВНЫЕ ОБОЗНАЧЕНИЯ"
        ws_legend[f'A{current_row}'].font = legend_header_font
        ws_legend[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws_legend[f'A{current_row}'].fill = legend_header_fill
        ws_legend[f'A{current_row}'].border = thick_border
        ws_legend.row_dimensions[current_row].height = 25
        current_row += 1

        legend_items = [
            ('Т', 'Теоретическое обучение', activity_fills.get('Т')),
            ('Э', 'Экзаменационная сессия', activity_fills.get('Э')),
            ('П', 'Практика (производственная, преддипломная)', activity_fills.get('П')),
            ('У', 'Учебная практика', activity_fills.get('У')),
            ('ПА', 'Промежуточная аттестация', activity_fills.get('ПА')),
            ('ГИА', 'Государственная итоговая аттестация', activity_fills.get('ГИА')),
            ('Г', 'Подготовка к сдаче и сдача гос. экзамена', activity_fills.get('Г')),
            ('Д', 'Подготовка и защита выпускной квалификационной работы', activity_fills.get('Д')),
            ('К', 'Каникулы', activity_fills.get('К')),
            ('*', 'Нерабочие праздничные дни', holiday_fill),
        ]

        for symbol, description, fill in legend_items:
            ws_legend[f'A{current_row}'] = symbol
            ws_legend[f'A{current_row}'].font = Font(name='Calibri', size=11, bold=True, color='000000')
            ws_legend[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws_legend[f'A{current_row}'].fill = fill
            ws_legend[f'A{current_row}'].border = thin_border

            ws_legend.merge_cells(f'B{current_row}:F{current_row}')
            ws_legend[f'B{current_row}'] = description
            ws_legend[f'B{current_row}'].font = legend_font
            ws_legend[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws_legend[f'B{current_row}'].border = thin_border

            ws_legend.row_dimensions[current_row].height = 22
            current_row += 1

        ws_legend.column_dimensions['A'].width = 8
        ws_legend.column_dimensions['B'].width = 50
        ws_legend.column_dimensions['C'].width = 10
        ws_legend.column_dimensions['D'].width = 10
        ws_legend.column_dimensions['E'].width = 10
        ws_legend.column_dimensions['F'].width = 10

        # ===== ЛИСТ 2: КАЛЕНДАРНЫЙ ГРАФИК =====
        ws = wb.create_sheet("Календарный график")

        current_row = 1

        ws.merge_cells(f'A{current_row}:BB{current_row}')
        ws[f'A{current_row}'] = f"КАЛЕНДАРНЫЙ УЧЕБНЫЙ ГРАФИК {start_year}-{start_year + program_years} г."
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{current_row}'].fill = legend_header_fill
        ws[f'A{current_row}'].border = thick_border
        ws.row_dimensions[current_row].height = 35
        current_row += 1

        current_row += 1

        for academic_year in range(program_years):
            actual_year = start_year + academic_year

            ws.merge_cells(f'A{current_row}:BB{current_row}')
            ws[f'A{current_row}'] = f"УЧЕБНЫЙ ГОД {actual_year}-{actual_year + 1}"
            ws[f'A{current_row}'].font = year_title_font
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'A{current_row}'].fill = year_header_fill
            ws[f'A{current_row}'].border = thick_border
            ws.row_dimensions[current_row].height = 28
            current_row += 1

            current_row += 1

            current_row = self.create_horizontal_calendar(
                ws, actual_year, generated_schedule,
                activity_fills, weekend_fill, holiday_fill,
                thin_border, header_font, header_fill,
                month_fill, month_fill_alt, data_font, current_row
            )

            current_row += 2

        ws.column_dimensions['A'].width = 6
        for col_idx in range(2, 60):
            ws.column_dimensions[get_column_letter(col_idx)].width = 4.5

        # ===== ЛИСТ 3: ИТОГИ =====
        self.create_beautiful_summary_sheet(wb, generated_schedule, start_year, program_type)

        return wb

    def create_horizontal_calendar(self, ws, start_year, generated_schedule,
                                   activity_fills, weekend_fill, holiday_fill,
                                   thin_border, header_font, header_fill,
                                   month_fill, month_fill_alt, data_font, start_row):
        """ГОРИЗОНТАЛЬНЫЙ КАЛЕНДАРЬ"""
        current_row = start_row

        start_date = datetime(start_year, 9, 1)
        end_date = datetime(start_year + 1, 8, 31)
        first_monday = self.get_monday_of_week(start_date)

        all_weeks = []
        current_date = first_monday
        while current_date <= end_date:
            week_dates = [current_date + timedelta(days=i) for i in range(7)]
            all_weeks.append(week_dates)
            current_date += timedelta(days=7)

        academic_months_list = [(start_year, m) for m in range(9, 13)] + \
                               [(start_year + 1, m) for m in range(1, 9)]
        month_to_index = {}
        for idx, month_key in enumerate(academic_months_list):
            month_to_index[month_key] = idx

        ws[f'A{current_row}'] = 'Месяц'
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = header_fill
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{current_row}'].border = thin_border

        month_columns = {}
        for week_idx, week_dates in enumerate(all_weeks):
            col = week_idx + 2
            days_in_range = [d for d in week_dates if start_date <= d <= end_date]
            if days_in_range:
                representative_date = days_in_range[-1]
                month_key = (representative_date.year, representative_date.month)
                if month_key not in month_columns:
                    month_columns[month_key] = []
                month_columns[month_key].append(col)

        academic_months = [(start_year, m) for m in range(9, 13)] + \
                          [(start_year + 1, m) for m in range(1, 9)]

        thick_month_border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='medium', color='B0BEC5'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )

        for month_idx, (year, month) in enumerate(academic_months):
            month_key = (year, month)
            if month_key in month_columns:
                cols = sorted(month_columns[month_key])
                start_col = cols[0]
                end_col = cols[-1]

                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=current_row, column=col)

                    if col == start_col:
                        cell.value = self.month_names_ru[month]
                        cell.font = Font(name='Calibri', size=10, bold=True, color='424242')

                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    if col == end_col:
                        cell.border = thick_month_border
                    else:
                        cell.border = thin_border

                if start_col != end_col:
                    ws.merge_cells(
                        f'{get_column_letter(start_col)}{current_row}:{get_column_letter(end_col)}{current_row}')

        ws.row_dimensions[current_row].height = 20
        current_row += 1

        days_of_week = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']

        for day_idx, day_name in enumerate(days_of_week):
            ws[f'A{current_row}'] = day_name
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].fill = header_fill
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'A{current_row}'].border = thin_border

            for week_idx, week_dates in enumerate(all_weeks):
                col = week_idx + 2
                date = week_dates[day_idx]

                cell = ws.cell(row=current_row, column=col)

                if start_date <= date <= end_date:
                    cell.value = date.day

                    month_key = (date.year, date.month)
                    month_index = month_to_index.get(month_key, 0)
                    is_bold = month_index % 2 == 0

                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    is_last_col_of_month = False
                    for m_key, cols in month_columns.items():
                        if col == max(cols):
                            is_last_col_of_month = True
                            break

                    cell.border = thick_month_border if is_last_col_of_month else thin_border

                    if self.is_holiday(date):
                        cell.fill = holiday_fill
                        cell.font = Font(name='Calibri', size=10, bold=True, color='D32F2F')
                    elif date.weekday() >= 5:
                        cell.fill = weekend_fill
                        cell.font = Font(name='Calibri', size=10, bold=is_bold, color='000000')
                    else:
                        cell.font = Font(name='Calibri', size=10, bold=is_bold, color='000000')
                else:
                    cell.value = ""
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    cell.border = thin_border

            ws.row_dimensions[current_row].height = 18
            current_row += 1

        ws[f'A{current_row}'] = 'Неделя'
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = header_fill
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{current_row}'].border = thin_border

        for week_idx in range(len(all_weeks)):
            col = week_idx + 2
            cell = ws.cell(row=current_row, column=col)
            cell.value = week_idx + 1
            cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="66BB6A", end_color="66BB6A", fill_type="solid")

            is_last_col_of_month = False
            for month_key, cols in month_columns.items():
                if col == max(cols):
                    is_last_col_of_month = True
                    break

            cell.border = thick_month_border if is_last_col_of_month else thin_border

        ws.row_dimensions[current_row].height = 20
        current_row += 1

        current_row += 1

        ws[f'A{current_row}'] = 'Занятия'
        ws[f'A{current_row}'].font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        ws[f'A{current_row}'].fill = PatternFill(start_color="66BB6A", end_color="66BB6A", fill_type="solid")
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{current_row}'].border = thin_border
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        for day_idx, day_name in enumerate(days_of_week):
            ws[f'A{current_row}'] = day_name
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].fill = header_fill
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'A{current_row}'].border = thin_border

            for week_idx, week_dates in enumerate(all_weeks):
                col = week_idx + 2
                date = week_dates[day_idx]

                cell = ws.cell(row=current_row, column=col)

                is_last_col_of_month = False
                for month_key, cols in month_columns.items():
                    if col == max(cols):
                        is_last_col_of_month = True
                        break

                thick_month_border_activity = Border(
                    left=Side(style='thin', color='E0E0E0'),
                    right=Side(style='medium', color='B0BEC5'),
                    top=Side(style='thin', color='E0E0E0'),
                    bottom=Side(style='thin', color='E0E0E0')
                )

                cell.border = thick_month_border_activity if is_last_col_of_month else thin_border

                if start_date <= date <= end_date:
                    if self.is_holiday(date):
                        cell.value = '*'
                        cell.fill = holiday_fill
                        cell.font = Font(name='Calibri', size=10, bold=True, color='D32F2F')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif date.weekday() >= 5:
                        cell.fill = weekend_fill
                    else:
                        activity_type = self.get_activity_for_date(date, generated_schedule)
                        if activity_type and activity_type in activity_fills:
                            cell.value = activity_type

                            month_key = (date.year, date.month)
                            month_index = month_to_index.get(month_key, 0)
                            is_bold = month_index % 2 == 0

                            cell.font = Font(name='Calibri', size=10, bold=is_bold, color='000000')
                            cell.fill = activity_fills[activity_type]
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.value = ""
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            ws.row_dimensions[current_row].height = 18
            current_row += 1

        return current_row

    def get_activity_for_date(self, date, generated_schedule):
        """Получить тип занятия для даты"""
        for period in generated_schedule:
            if date in period['days']:
                return period['type']
        return None

    def create_beautiful_summary_sheet(self, wb, generated_schedule, start_year, program_type):
        """Создать итоговую таблицу точно как в примере"""
        ws = wb.create_sheet("Итоги")

        program_years = 2 if "Ординатура" in program_type else 3

        # Стили
        title_font = Font(name='Calibri', size=12, bold=True, color='1976D2')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        bold_font = Font(name='Calibri', size=10, bold=True, color='000000')
        data_font = Font(name='Calibri', size=10, color='000000')

        header_fill = PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")

        # Собираем статистику
        stats = {}
        holidays_count = {}  # Подсчет праздничных дней

        for period in generated_schedule:
            year = period['year']
            semester = period['semester']
            activity_type = period['type']
            weeks = period['weeks']
            days = len(period['days'])

            key = (year, semester)
            if key not in stats:
                stats[key] = {}
                holidays_count[key] = 0

            # Подсчет праздничных дней
            for day in period['days']:
                if self.is_holiday(day):
                    holidays_count[key] += 1

            # Обработка ГИА
            if activity_type == 'ГИА':
                if 'Г' not in stats[key]:
                    stats[key]['Г'] = {'weeks': 0, 'days': 0}
                if 'Д' not in stats[key]:
                    stats[key]['Д'] = {'weeks': 0, 'days': 0}
                stats[key]['Г']['weeks'] += weeks / 2
                stats[key]['Г']['days'] += days // 2
                stats[key]['Д']['weeks'] += weeks / 2
                stats[key]['Д']['days'] += days - (days // 2)
                continue

            if activity_type not in stats[key]:
                stats[key][activity_type] = {'weeks': 0, 'days': 0}

            stats[key][activity_type]['weeks'] += weeks
            stats[key][activity_type]['days'] += days

        current_row = 1

        # ===== ЗАГОЛОВОК "Сводные данные" =====
        ws[f'A{current_row}'] = 'Сводные данные'
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[current_row].height = 25
        current_row += 2

        # ===== ПУСТЫЕ СТРОКИ ДЛЯ ЗАГОЛОВКОВ =====
        current_row += 1

        # ===== СТРОКА С КУРСАМИ =====
        col_offset = 4  # ПРИБЛИЖАЕМ К НАЗВАНИЯМ! Колонка D вместо O

        for year in range(1, program_years + 1):
            col_start = col_offset + (year - 1) * 6
            col_end = col_start + 5
            ws.merge_cells(f'{get_column_letter(col_start)}{current_row}:{get_column_letter(col_end)}{current_row}')
            cell = ws[f'{get_column_letter(col_start)}{current_row}']
            cell.value = f'Курс {year}'
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        current_row += 1

        # ===== СТРОКА С СЕМЕСТРАМИ =====
        for year in range(1, program_years + 1):
            col_start = col_offset + (year - 1) * 6
            sem_base = (year - 1) * 2

            # Сем. X
            col1 = col_start
            col2 = col_start + 1
            ws.merge_cells(f'{get_column_letter(col1)}{current_row}:{get_column_letter(col2)}{current_row}')
            cell = ws[f'{get_column_letter(col1)}{current_row}']
            cell.value = f'Сем. {sem_base + 1}'
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Сем. Y
            col1 = col_start + 2
            col2 = col_start + 3
            ws.merge_cells(f'{get_column_letter(col1)}{current_row}:{get_column_letter(col2)}{current_row}')
            cell = ws[f'{get_column_letter(col1)}{current_row}']
            cell.value = f'Сем. {sem_base + 2}'
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Всего
            col1 = col_start + 4
            col2 = col_start + 5
            ws.merge_cells(f'{get_column_letter(col1)}{current_row}:{get_column_letter(col2)}{current_row}')
            cell = ws[f'{get_column_letter(col1)}{current_row}']
            cell.value = 'Всего'
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        current_row += 1

        # ===== ДАННЫЕ =====
        def format_weeks(weeks):
            if weeks < 0.01:
                return ''
            whole = int(weeks)
            fraction = weeks - whole
            if fraction < 0.01:
                return str(whole) if whole > 0 else ''
            sixths = round(fraction * 6)
            if sixths == 0:
                return str(whole) if whole > 0 else ''
            elif sixths == 6:
                return str(whole + 1)
            else:
                return f'{whole} {sixths}/6' if whole > 0 else f'{sixths}/6'

        # Список типов деятельности в правильном порядке
        activity_types = [
            ('', 'Теоретическое обучение', 'Т'),
            ('Э', 'Экзаменационные сессии', 'Э'),
            ('У', 'Учебная практика', 'У'),
            ('Н', 'Научно-исслед. работа', 'Н'),
            ('П', 'Производственная практика', 'П'),
            ('Пд', 'Преддипломная практика', 'Пд'),
            ('ПА', 'Повторная, вторая повторная промежуточная аттестация', 'ПА'),
            ('Д', 'Подготовка к процедуре защиты и защита выпускной квалификационной работы', 'Д'),
            ('Г', 'Подготовка к сдаче и сдача гос. экзамена', 'Г'),
            ('К', 'Каникулы', 'К'),
            ('*', 'Нерабочие праздничные дни (не включая воскресенья)', '*'),
        ]

        for symbol, name, code in activity_types:
            # Символ (колонка A)
            ws[f'A{current_row}'] = symbol
            ws[f'A{current_row}'].font = bold_font
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')

            # Название (колонка B)
            ws[f'B{current_row}'] = name
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')

            # Данные
            for year in range(1, program_years + 1):
                col_start = col_offset + (year - 1) * 6

                # Семестр 1
                key = (year, 1)
                sem1_weeks = 0
                sem1_days = 0
                if code == '*':
                    sem1_days = holidays_count.get(key, 0)
                elif key in stats and code in stats[key]:
                    sem1_weeks = stats[key][code]['weeks']
                    sem1_days = stats[key][code]['days']

                col1 = col_start
                col2 = col_start + 1
                ws.merge_cells(f'{get_column_letter(col1)}{current_row}:{get_column_letter(col2)}{current_row}')
                cell = ws[f'{get_column_letter(col1)}{current_row}']
                if code == '*':
                    cell.value = f'{sem1_days} дн' if sem1_days > 0 else ''
                else:
                    cell.value = format_weeks(sem1_weeks)
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Семестр 2
                key = (year, 2)
                sem2_weeks = 0
                sem2_days = 0
                if code == '*':
                    sem2_days = holidays_count.get(key, 0)
                elif key in stats and code in stats[key]:
                    sem2_weeks = stats[key][code]['weeks']
                    sem2_days = stats[key][code]['days']

                col1 = col_start + 2
                col2 = col_start + 3
                ws.merge_cells(f'{get_column_letter(col1)}{current_row}:{get_column_letter(col2)}{current_row}')
                cell = ws[f'{get_column_letter(col1)}{current_row}']
                if code == '*':
                    cell.value = f'{sem2_days} дн' if sem2_days > 0 else ''
                else:
                    cell.value = format_weeks(sem2_weeks)
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Всего за год
                year_total_weeks = sem1_weeks + sem2_weeks
                year_total_days = sem1_days + sem2_days
                col1 = col_start + 4
                col2 = col_start + 5
                ws.merge_cells(f'{get_column_letter(col1)}{current_row}:{get_column_letter(col2)}{current_row}')
                cell = ws[f'{get_column_letter(col1)}{current_row}']
                if code == '*':
                    cell.value = f'{year_total_days} дн' if year_total_days > 0 else ''
                else:
                    cell.value = format_weeks(year_total_weeks)
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            ws.row_dimensions[current_row].height = 25
            current_row += 1

            # Дополнительная строка для экзаменов (дни)
            if code == 'Э':
                ws[f'A{current_row}'] = symbol
                ws[f'A{current_row}'].font = bold_font
                ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')

                ws[f'B{current_row}'] = name
                ws[f'B{current_row}'].font = data_font
                ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')

                for year in range(1, program_years + 1):
                    col_start = col_offset + (year - 1) * 6

                    # Сем 1 дни
                    key = (year, 1)
                    days1 = stats[key][code]['days'] if key in stats and code in stats[key] else 0
                    col1 = col_start
                    col2 = col_start + 1
                    ws.merge_cells(f'{get_column_letter(col1)}{current_row}:{get_column_letter(col2)}{current_row}')
                    cell = ws[f'{get_column_letter(col1)}{current_row}']
                    cell.value = f'{days1} дн' if days1 > 0 else ''
                    cell.font = data_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    # Сем 2 дни
                    key = (year, 2)
                    days2 = stats[key][code]['days'] if key in stats and code in stats[key] else 0
                    col1 = col_start + 2
                    col2 = col_start + 3
                    ws.merge_cells(f'{get_column_letter(col1)}{current_row}:{get_column_letter(col2)}{current_row}')
                    cell = ws[f'{get_column_letter(col1)}{current_row}']
                    cell.value = f'{days2} дн' if days2 > 0 else ''
                    cell.font = data_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    # Всего дней
                    total_days = days1 + days2
                    col1 = col_start + 4
                    col2 = col_start + 5
                    ws.merge_cells(f'{get_column_letter(col1)}{current_row}:{get_column_letter(col2)}{current_row}')
                    cell = ws[f'{get_column_letter(col1)}{current_row}']
                    cell.value = f'{total_days} дн' if total_days > 0 else ''
                    cell.font = data_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                ws.row_dimensions[current_row].height = 25
                current_row += 1

            # Дополнительная строка для каникул (Продолжительность каникул)
            elif code == 'К':
                ws[f'A{current_row}'] = symbol
                ws[f'A{current_row}'].font = bold_font
                ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')

                ws[f'B{current_row}'] = 'Продолжительность каникул'
                ws[f'B{current_row}'].font = data_font
                ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')

                for year in range(1, program_years + 1):
                    col_start = col_offset + (year - 1) * 6

                    # Сем 1 дни
                    key = (year, 1)
                    days1 = stats[key][code]['days'] if key in stats and code in stats[key] else 0
                    col1 = col_start
                    col2 = col_start + 1
                    ws.merge_cells(f'{get_column_letter(col1)}{current_row}:{get_column_letter(col2)}{current_row}')
                    cell = ws[f'{get_column_letter(col1)}{current_row}']
                    cell.value = f'{days1} дн' if days1 > 0 else ''
                    cell.font = data_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    # Сем 2 дни
                    key = (year, 2)
                    days2 = stats[key][code]['days'] if key in stats and code in stats[key] else 0
                    col1 = col_start + 2
                    col2 = col_start + 3
                    ws.merge_cells(f'{get_column_letter(col1)}{current_row}:{get_column_letter(col2)}{current_row}')
                    cell = ws[f'{get_column_letter(col1)}{current_row}']
                    cell.value = f'{days2} дн' if days2 > 0 else ''
                    cell.font = data_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    # Всего дней
                    total_days = days1 + days2
                    col1 = col_start + 4
                    col2 = col_start + 5
                    ws.merge_cells(f'{get_column_letter(col1)}{current_row}:{get_column_letter(col2)}{current_row}')
                    cell = ws[f'{get_column_letter(col1)}{current_row}']
                    cell.value = f'{total_days} дн' if total_days > 0 else ''
                    cell.font = data_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                ws.row_dimensions[current_row].height = 25
                current_row += 1

            # Строка "в том числе ДО"
            ws[f'B{current_row}'] = 'в том числе ДО'
            ws[f'B{current_row}'].font = data_font
            ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            current_row += 1

        # ===== ПУСТАЯ СТРОКА =====
        current_row += 1

        # ===== ПРОДОЛЖИТЕЛЬНОСТЬ ОБУЧЕНИЯ =====
        ws[f'A{current_row}'] = 'Продолжительность обучения '
        ws[f'A{current_row}'].font = bold_font

        for year in range(1, program_years + 1):
            col_start = col_offset + (year - 1) * 6
            col_end = col_start + 5
            ws.merge_cells(f'{get_column_letter(col_start)}{current_row}:{get_column_letter(col_end)}{current_row}')
            cell = ws[f'{get_column_letter(col_start)}{current_row}']
            cell.value = 'более 39 нед.'
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        current_row += 1

        # ===== ИТОГО =====
        ws[f'A{current_row}'] = ' Итого'
        ws[f'A{current_row}'].font = bold_font

        for year in range(1, program_years + 1):
            col_start = col_offset + (year - 1) * 6

            # Сем 1
            key = (year, 1)
            sem1_total = 0
            if key in stats:
                for act_type in stats[key]:
                    sem1_total += stats[key][act_type]['weeks']

            col1 = col_start
            col2 = col_start + 1
            ws.merge_cells(f'{get_column_letter(col1)}{current_row}:{get_column_letter(col2)}{current_row}')
            cell = ws[f'{get_column_letter(col1)}{current_row}']
            cell.value = format_weeks(sem1_total)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Сем 2
            key = (year, 2)
            sem2_total = 0
            if key in stats:
                for act_type in stats[key]:
                    sem2_total += stats[key][act_type]['weeks']

            col1 = col_start + 2
            col2 = col_start + 3
            ws.merge_cells(f'{get_column_letter(col1)}{current_row}:{get_column_letter(col2)}{current_row}')
            cell = ws[f'{get_column_letter(col1)}{current_row}']
            cell.value = format_weeks(sem2_total)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Всего
            year_total = sem1_total + sem2_total
            col1 = col_start + 4
            col2 = col_start + 5
            ws.merge_cells(f'{get_column_letter(col1)}{current_row}:{get_column_letter(col2)}{current_row}')
            cell = ws[f'{get_column_letter(col1)}{current_row}']
            cell.value = format_weeks(year_total)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        current_row += 1

        # ===== ПРОДОЛЖИТЕЛЬНОСТЬ =====
        ws[f'A{current_row}'] = ' Продолжительность'
        ws[f'A{current_row}'].font = bold_font

        for year in range(1, program_years + 1):
            col_start = col_offset + (year - 1) * 6

            # Сем 1
            key = (year, 1)
            sem1_days = 0
            if key in stats:
                for act_type in stats[key]:
                    sem1_days += stats[key][act_type]['days']

            col1 = col_start
            col2 = col_start + 1
            ws.merge_cells(f'{get_column_letter(col1)}{current_row}:{get_column_letter(col2)}{current_row}')
            cell = ws[f'{get_column_letter(col1)}{current_row}']
            cell.value = f'{sem1_days} дн' if sem1_days > 0 else ''
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Сем 2
            key = (year, 2)
            sem2_days = 0
            if key in stats:
                for act_type in stats[key]:
                    sem2_days += stats[key][act_type]['days']

            col1 = col_start + 2
            col2 = col_start + 3
            ws.merge_cells(f'{get_column_letter(col1)}{current_row}:{get_column_letter(col2)}{current_row}')
            cell = ws[f'{get_column_letter(col1)}{current_row}']
            cell.value = f'{sem2_days} дн' if sem2_days > 0 else ''
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Всего
            year_days = sem1_days + sem2_days
            col1 = col_start + 4
            col2 = col_start + 5
            ws.merge_cells(f'{get_column_letter(col1)}{current_row}:{get_column_letter(col2)}{current_row}')
            cell = ws[f'{get_column_letter(col1)}{current_row}']
            cell.value = f'{year_days} дн' if year_days > 0 else ''
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        current_row += 1

        # ===== ВИСОКОСНЫЙ ГОД =====
        ws[f'A{current_row}'] = ' Високосный год'
        ws[f'A{current_row}'].font = bold_font

        for year in range(1, program_years + 1):
            col_start = col_offset + (year - 1) * 6
            col_end = col_start + 5
            ws.merge_cells(f'{get_column_letter(col_start)}{current_row}:{get_column_letter(col_end)}{current_row}')
            cell = ws[f'{get_column_letter(col_start)}{current_row}']
            cell.value = '-'
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        current_row += 1

        # ===== СТУДЕНТОВ =====
        ws[f'A{current_row}'] = ' Студентов'
        ws[f'A{current_row}'].font = bold_font
        current_row += 1

        # Пустые строки
        current_row += 3

        # ===== ГРУПП =====
        ws[f'A{current_row}'] = ' Групп'
        ws[f'A{current_row}'].font = bold_font

        # Настройка ширины колонок
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 70
        ws.column_dimensions['C'].width = 2
        for col_idx in range(4, 30):
            ws.column_dimensions[get_column_letter(col_idx)].width = 10


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.app = EducationalScheduleApp()
        self.periods_data = []
        self.generated_schedule = None
        self.start_year = 2025
        self.program_type = "Ординатура (2 года)"

        self.init_ui()
        self.apply_styles()

    def init_ui(self):
        self.setWindowTitle('Учебный график - Итоги как в примере')
        self.setGeometry(100, 100, 1500, 900)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(0)
        main_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)

        scroll_content = QWidget()
        scroll_content.setStyleSheet("background-color: #0e1117;")
        content_layout = QVBoxLayout(scroll_content)
        content_layout.setSpacing(0)
        content_layout.setContentsMargins(0, 0, 0, 0)

        content_container = QWidget()
        content_container.setStyleSheet("background-color: #0e1117;")
        container_layout = QVBoxLayout(content_container)
        container_layout.setContentsMargins(50, 40, 50, 50)
        container_layout.setSpacing(32)

        header_row = QHBoxLayout()
        header_row.setSpacing(24)

        title_layout = QVBoxLayout()
        title_layout.setSpacing(8)

        title = QLabel('📚 Учебный график')
        title.setObjectName("mainTitle")

        subtitle = QLabel('✅ Итоги в формате примера - точная копия!')
        subtitle.setObjectName("subtitle")

        title_layout.addWidget(title)
        title_layout.addWidget(subtitle)

        header_row.addLayout(title_layout)
        header_row.addStretch()

        container_layout.addLayout(header_row)

        header_line = QFrame()
        header_line.setFrameShape(QFrame.Shape.HLine)
        header_line.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 transparent, stop:0.2 #34d399, stop:0.8 #34d399, stop:1 transparent);
                border: none;
                height: 3px;
                margin: 16px 0px;
            }
        """)
        container_layout.addWidget(header_line)
        container_layout.addSpacing(8)

        settings_row = QHBoxLayout()
        settings_row.setSpacing(20)

        program_layout = QVBoxLayout()
        program_layout.setSpacing(8)
        program_label = QLabel('Тип программы')
        program_label.setObjectName("inputLabel")
        self.program_combo = QComboBox()
        self.program_combo.addItems(['Ординатура (2 года)', 'Аспирантура (3 года)'])
        self.program_combo.currentTextChanged.connect(self.on_program_changed)
        program_layout.addWidget(program_label)
        program_layout.addWidget(self.program_combo)

        year_layout = QVBoxLayout()
        year_layout.setSpacing(8)
        year_label = QLabel('Начальный год')
        year_label.setObjectName("inputLabel")
        self.year_combo = QComboBox()
        self.year_combo.addItems(['2025', '2026', '2027'])
        self.year_combo.currentTextChanged.connect(self.on_year_changed)
        year_layout.addWidget(year_label)
        year_layout.addWidget(self.year_combo)

        settings_row.addLayout(program_layout)
        settings_row.addLayout(year_layout)
        settings_row.addStretch()

        container_layout.addLayout(settings_row)

        button_row = QHBoxLayout()
        button_row.setSpacing(12)

        example_btn = QPushButton('📋 Загрузить пример')
        example_btn.setObjectName("secondaryButton")
        example_btn.clicked.connect(self.load_example)

        clear_btn = QPushButton('🗑️ Очистить')
        clear_btn.setObjectName("secondaryButton")
        clear_btn.clicked.connect(self.clear_data)

        button_row.addWidget(example_btn)
        button_row.addWidget(clear_btn)
        button_row.addStretch()

        container_layout.addLayout(button_row)
        container_layout.addSpacing(16)

        periods_label = QLabel('Периоды обучения')
        periods_label.setObjectName("sectionTitle")
        container_layout.addWidget(periods_label)

        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(['Год', 'Семестр', 'Тип', 'Недели'])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.verticalHeader().setDefaultSectionSize(82)
        self.table.setAlternatingRowColors(True)
        self.table.setShowGrid(False)
        self.table.setMinimumHeight(400)
        self.table.itemChanged.connect(self.update_weeks_total)
        container_layout.addWidget(self.table)

        # Метка для отображения суммы недель
        self.weeks_total_label = QLabel('Всего недель: 0')
        self.weeks_total_label.setObjectName("weeksTotalLabel")
        container_layout.addWidget(self.weeks_total_label)

        table_btn_row = QHBoxLayout()
        table_btn_row.setSpacing(12)

        add_row_btn = QPushButton('➕ Добавить строку')
        add_row_btn.setObjectName("secondaryButton")
        add_row_btn.clicked.connect(self.add_row)

        remove_row_btn = QPushButton('➖ Удалить строку')
        remove_row_btn.setObjectName("secondaryButton")
        remove_row_btn.clicked.connect(self.remove_row)

        table_btn_row.addWidget(add_row_btn)
        table_btn_row.addWidget(remove_row_btn)
        table_btn_row.addStretch()

        container_layout.addLayout(table_btn_row)
        container_layout.addSpacing(16)

        action_row = QHBoxLayout()
        action_row.setSpacing(16)

        generate_btn = QPushButton('🎓 Сгенерировать график')
        generate_btn.setObjectName("primaryButton")
        generate_btn.clicked.connect(self.generate_schedule)

        self.download_btn = QPushButton('📥 Скачать Excel')
        self.download_btn.setObjectName("downloadButton")
        self.download_btn.clicked.connect(self.download_excel)
        self.download_btn.setEnabled(False)

        action_row.addWidget(generate_btn, 1)
        action_row.addWidget(self.download_btn, 1)

        container_layout.addLayout(action_row)

        self.preview_section = QWidget()
        self.preview_section.setStyleSheet("background-color: #0e1117;")
        preview_layout = QVBoxLayout(self.preview_section)
        preview_layout.setContentsMargins(0, 32, 0, 0)
        preview_layout.setSpacing(16)

        preview_label = QLabel('Предварительный просмотр')
        preview_label.setObjectName("sectionTitle")
        preview_layout.addWidget(preview_label)

        self.preview_table = QTableWidget()
        self.preview_table.setColumnCount(7)
        self.preview_table.setHorizontalHeaderLabels(['Год', 'Семестр', 'Тип', 'Недели', 'Начало', 'Конец', 'Дней'])
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.preview_table.verticalHeader().setVisible(False)
        self.preview_table.verticalHeader().setDefaultSectionSize(60)
        self.preview_table.setAlternatingRowColors(True)
        self.preview_table.setShowGrid(False)
        self.preview_table.setMinimumHeight(350)
        preview_layout.addWidget(self.preview_table)

        self.preview_section.setVisible(False)
        container_layout.addWidget(self.preview_section)

        footer_layout = QVBoxLayout()
        footer_layout.setContentsMargins(0, 32, 0, 0)
        footer_layout.setSpacing(0)

        authors_label = QLabel('Разработчики: Бахмутов Е., Клюев П. | v4.1 - ТОЧНАЯ КОПИЯ ПРИМЕРА!')
        authors_label.setObjectName("authorsLabel")
        authors_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        footer_layout.addWidget(authors_label)

        container_layout.addLayout(footer_layout)

        content_layout.addWidget(content_container)

        scroll.setWidget(scroll_content)
        main_layout.addWidget(scroll)

    def apply_styles(self):
        self.setStyleSheet("""
            QMainWindow {
                background-color: #0e1117;
            }

            QLabel#mainTitle {
                font-size: 42px;
                font-weight: 800;
                color: #fafafa;
                margin: 0;
                padding: 0;
                letter-spacing: -0.5px;
            }

            QLabel#subtitle {
                font-size: 18px;
                color: #34d399;
                margin-top: 4px;
                font-weight: 500;
                letter-spacing: 0.3px;
            }

            QLabel#sectionTitle {
                font-size: 20px;
                font-weight: 600;
                color: #fafafa;
                margin-bottom: 8px;
            }

            QLabel#inputLabel {
                font-size: 14px;
                font-weight: 600;
                color: #fafafa;
            }

            QLabel#authorsLabel {
                font-size: 13px;
                color: #6b7280;
                font-weight: 400;
                opacity: 0.7;
            }

            QLabel#weeksTotalLabel {
                font-size: 16px;
                font-weight: 600;
                color: #34d399;
                margin-top: 8px;
                padding: 8px;
                background-color: #1a1c24;
                border-radius: 6px;
                border: 1px solid #31343f;
            }

            QComboBox {
                padding: 12px 16px;
                border: 1px solid #464a5e;
                border-radius: 8px;
                background-color: #262730;
                font-size: 15px;
                min-width: 240px;
                color: #fafafa;
                min-height: 44px;
                font-weight: 400;
            }

            QComboBox:hover {
                border-color: #2d8659;
            }

            QComboBox:focus {
                border-color: #2d8659;
                outline: none;
            }

            QComboBox::drop-down {
                border: none;
                width: 32px;
            }

            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 7px solid #a3a8b4;
                margin-right: 10px;
            }

            QComboBox QAbstractItemView {
                background-color: #262730;
                border: 1px solid #464a5e;
                selection-background-color: #2d8659;
                selection-color: #ffffff;
                outline: none;
                padding: 6px;
                font-size: 15px;
                color: #fafafa;
            }

            QComboBox QAbstractItemView::item {
                padding: 10px 12px;
                min-height: 36px;
                color: #fafafa;
            }

            QComboBox QAbstractItemView::item:hover {
                background-color: #2d8659;
                color: #ffffff;
            }

            QPushButton#primaryButton {
                padding: 14px 32px;
                border: 2px solid #34d399;
                border-radius: 10px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #34d399, stop:0.5 #2d8659, stop:1 #10b981);
                color: white;
                font-size: 17px;
                font-weight: 700;
                min-height: 56px;
                letter-spacing: 1px;
            }

            QPushButton#primaryButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #6ee7b7, stop:0.5 #34d399, stop:1 #10b981);
                border-color: #6ee7b7;
                border-width: 3px;
                padding: 13px 31px;
            }

            QPushButton#primaryButton:pressed {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #22c55e, stop:0.5 #16a34a, stop:1 #15803d);
                border-color: #16a34a;
                border-width: 2px;
                padding: 14px 32px;
            }

            QPushButton#downloadButton {
                padding: 14px 32px;
                border: 2px solid #3b82f6;
                border-radius: 10px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3b82f6, stop:0.5 #2563eb, stop:1 #1d4ed8);
                color: white;
                font-size: 17px;
                font-weight: 700;
                min-height: 56px;
                letter-spacing: 0.5px;
            }

            QPushButton#downloadButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #93c5fd, stop:0.5 #60a5fa, stop:1 #3b82f6);
                border-color: #93c5fd;
                border-width: 3px;
                padding: 13px 31px;
            }

            QPushButton#downloadButton:pressed {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #2563eb, stop:0.5 #1d4ed8, stop:1 #1e40af);
                border-color: #1d4ed8;
                border-width: 2px;
                padding: 14px 32px;
            }

            QPushButton#downloadButton:disabled {
                background: #1a1c24;
                color: #464a5e;
                border-color: #31343f;
                border-width: 1px;
            }

            QPushButton#secondaryButton {
                padding: 10px 20px;
                border: 2px solid #464a5e;
                border-radius: 8px;
                background-color: #262730;
                color: #fafafa;
                font-size: 15px;
                font-weight: 500;
                min-height: 40px;
            }

            QPushButton#secondaryButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #3d4150, stop:1 #2d3038);
                border-color: #34d399;
                border-width: 2px;
                color: #34d399;
            }

            QPushButton#secondaryButton:pressed {
                background-color: #1c1f26;
                border-color: #2d8659;
                color: #2d8659;
            }

            QTableWidget {
                border: 1px solid #31343f;
                border-radius: 8px;
                background-color: #1a1c24;
                font-size: 16px;
                color: #fafafa;
            }

            QTableWidget::item {
                padding: 16px;
                color: #fafafa;
                background-color: #1a1c24;
                font-size: 16px;
                border: none;
            }

            QTableWidget::item:selected {
                background-color: #262730;
                color: #fafafa;
            }

            QHeaderView::section {
                background-color: #262730;
                padding: 16px;
                border: none;
                border-bottom: 2px solid #31343f;
                font-weight: 600;
                font-size: 15px;
                color: #fafafa;
            }

            QTableWidget::item:alternate {
                background-color: #14161d;
            }

            QScrollArea {
                border: none;
                background-color: #0e1117;
            }

            QScrollBar:vertical {
                border: none;
                background: #1a1c24;
                width: 12px;
                margin: 0px;
            }

            QScrollBar::handle:vertical {
                background: #464a5e;
                border-radius: 6px;
                min-height: 30px;
            }

            QScrollBar::handle:vertical:hover {
                background: #5a5f75;
            }

            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
            }
        """)

    def on_program_changed(self, text):
        self.program_type = text

    def on_year_changed(self, text):
        self.start_year = int(text)

    def load_example(self):
        if "Аспирантура" in self.program_type:
            self.periods_data = [
                {"Год": 1, "Семестр": 1, "Тип": "Т", "Недели": 12},
                {"Год": 1, "Семестр": 1, "Тип": "Э", "Недели": 2},
                {"Год": 1, "Семестр": 1, "Тип": "П", "Недели": 8},
                {"Год": 1, "Семестр": 1, "Тип": "ПА", "Недели": 1},
                {"Год": 1, "Семестр": 2, "Тип": "Т", "Недели": 10},
                {"Год": 1, "Семестр": 2, "Тип": "Э", "Недели": 2},
                {"Год": 1, "Семестр": 2, "Тип": "П", "Недели": 6},
                {"Год": 1, "Семестр": 2, "Тип": "ПА", "Недели": 1},
                {"Год": 1, "Семестр": 2, "Тип": "К", "Недели": 6},

                {"Год": 2, "Семестр": 1, "Тип": "Т", "Недели": 12},
                {"Год": 2, "Семестр": 1, "Тип": "Э", "Недели": 2},
                {"Год": 2, "Семестр": 1, "Тип": "П", "Недели": 8},
                {"Год": 2, "Семестр": 1, "Тип": "ПА", "Недели": 1},
                {"Год": 2, "Семестр": 2, "Тип": "Т", "Недели": 10},
                {"Год": 2, "Семестр": 2, "Тип": "Э", "Недели": 2},
                {"Год": 2, "Семестр": 2, "Тип": "П", "Недели": 6},
                {"Год": 2, "Семестр": 2, "Тип": "ПА", "Недели": 1},
                {"Год": 2, "Семестр": 2, "Тип": "К", "Недели": 6},

                {"Год": 3, "Семестр": 1, "Тип": "Т", "Недели": 10},
                {"Год": 3, "Семестр": 1, "Тип": "У", "Недели": 4},
                {"Год": 3, "Семестр": 1, "Тип": "П", "Недели": 8},
                {"Год": 3, "Семестр": 1, "Тип": "ПА", "Недели": 1},
                {"Год": 3, "Семестр": 2, "Тип": "Т", "Недели": 6},
                {"Год": 3, "Семестр": 2, "Тип": "П", "Недели": 6},
                {"Год": 3, "Семестр": 2, "Тип": "Г", "Недели": 2},
                {"Год": 3, "Семестр": 2, "Тип": "Д", "Недели": 4},
                {"Год": 3, "Семестр": 2, "Тип": "К", "Недели": 8}
            ]
        else:
            self.periods_data = [
                {"Год": 1, "Семестр": 1, "Тип": "Т", "Недели": 10},
                {"Год": 1, "Семестр": 1, "Тип": "П", "Недели": 12},
                {"Год": 1, "Семестр": 1, "Тип": "ПА", "Недели": 1},
                {"Год": 1, "Семестр": 2, "Тип": "Т", "Недели": 4},
                {"Год": 1, "Семестр": 2, "Тип": "П", "Недели": 16},
                {"Год": 1, "Семестр": 2, "Тип": "ПА", "Недели": 1},
                {"Год": 1, "Семестр": 2, "Тип": "К", "Недели": 6},
                {"Год": 2, "Семестр": 1, "Тип": "Т", "Недели": 10},
                {"Год": 2, "Семестр": 1, "Тип": "П", "Недели": 12},
                {"Год": 2, "Семестр": 1, "Тип": "ПА", "Недели": 1},
                {"Год": 2, "Семестр": 2, "Тип": "Т", "Недели": 9},
                {"Год": 2, "Семестр": 2, "Тип": "П", "Недели": 8},
                {"Год": 2, "Семестр": 2, "Тип": "ПА", "Недели": 1},
                {"Год": 2, "Семестр": 2, "Тип": "ГИА", "Недели": 2},
                {"Год": 2, "Семестр": 2, "Тип": "К", "Недели": 6}
            ]
        self.update_table()

    def clear_data(self):
        self.periods_data = []
        self.update_table()
        self.preview_section.setVisible(False)
        self.download_btn.setEnabled(False)

    def add_row(self):
        row_position = self.table.rowCount()
        self.table.insertRow(row_position)

        table_combo_style = """
            QComboBox {
                padding: 11px 16px;
                border: 1px solid #464a5e;
                border-radius: 6px;
                background-color: #262730;
                font-size: 16px;
                color: #fafafa;
            }
            QComboBox:hover {
                border-color: #34d399;
                border-width: 2px;
                background-color: #2d3038;
            }
            QComboBox::drop-down {
                border: none;
                width: 28px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 7px solid #a3a8b4;
                margin-right: 10px;
            }
            QComboBox:hover::down-arrow {
                border-top-color: #34d399;
            }
            QComboBox QAbstractItemView {
                background-color: #262730;
                border: 1px solid #464a5e;
                selection-background-color: #2d8659;
                selection-color: #ffffff;
                font-size: 16px;
            }
            QComboBox QAbstractItemView::item {
                padding: 12px 14px;
                min-height: 38px;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: #34d399;
            }
        """

        year_combo = QComboBox()
        year_combo.addItems(['1', '2', '3'])
        year_combo.setStyleSheet(table_combo_style)
        year_container = QWidget()
        year_container.setStyleSheet("background-color: transparent;")
        year_layout = QHBoxLayout(year_container)
        year_layout.addWidget(year_combo)
        year_layout.setContentsMargins(6, 0, 6, 0)
        year_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.table.setCellWidget(row_position, 0, year_container)

        semester_combo = QComboBox()
        semester_combo.addItems(['1', '2'])
        semester_combo.setStyleSheet(table_combo_style)
        semester_container = QWidget()
        semester_container.setStyleSheet("background-color: transparent;")
        semester_layout = QHBoxLayout(semester_container)
        semester_layout.addWidget(semester_combo)
        semester_layout.setContentsMargins(6, 0, 6, 0)
        semester_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.table.setCellWidget(row_position, 1, semester_container)

        type_combo = QComboBox()
        type_combo.addItems(['Т', 'Э', 'П', 'У', 'ПА', 'ГИА', 'Г', 'Д', 'К'])
        type_combo.setStyleSheet(table_combo_style)
        type_container = QWidget()
        type_container.setStyleSheet("background-color: transparent;")
        type_layout = QHBoxLayout(type_container)
        type_layout.addWidget(type_combo)
        type_layout.setContentsMargins(6, 0, 6, 0)
        type_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.table.setCellWidget(row_position, 2, type_container)

        weeks_item = QTableWidgetItem('1.0')
        weeks_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
        font = QFont()
        font.setPointSize(16)
        weeks_item.setFont(font)
        self.table.setItem(row_position, 3, weeks_item)

        self.table.setRowHeight(row_position, 82)

    def remove_row(self):
        current_row = self.table.currentRow()
        if current_row >= 0:
            self.table.removeRow(current_row)
            self.update_weeks_total()

    def update_weeks_total(self):
        """Обновить подсчет общего количества недель"""
        total_weeks = 0
        for row in range(self.table.rowCount()):
            weeks_item = self.table.item(row, 3)
            if weeks_item:
                try:
                    weeks = float(weeks_item.text())
                    total_weeks += weeks
                except ValueError:
                    pass
        self.weeks_total_label.setText(f'Всего недель: {total_weeks:.1f}')

    def update_table(self):
        self.table.setRowCount(0)

        table_combo_style = """
            QComboBox {
                padding: 11px 16px;
                border: 1px solid #464a5e;
                border-radius: 6px;
                background-color: #262730;
                font-size: 16px;
                color: #fafafa;
            }
            QComboBox:hover {
                border-color: #34d399;
                border-width: 2px;
                background-color: #2d3038;
            }
            QComboBox::drop-down {
                border: none;
                width: 28px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 7px solid #a3a8b4;
                margin-right: 10px;
            }
            QComboBox:hover::down-arrow {
                border-top-color: #34d399;
            }
            QComboBox QAbstractItemView {
                background-color: #262730;
                border: 1px solid #464a5e;
                selection-background-color: #2d8659;
                selection-color: #ffffff;
                font-size: 16px;
            }
            QComboBox QAbstractItemView::item {
                padding: 12px 14px;
                min-height: 38px;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: #34d399;
            }
        """

        for data in self.periods_data:
            row_position = self.table.rowCount()
            self.table.insertRow(row_position)

            year_combo = QComboBox()
            year_combo.addItems(['1', '2', '3'])
            year_combo.setCurrentText(str(data['Год']))
            year_combo.setStyleSheet(table_combo_style)
            year_container = QWidget()
            year_container.setStyleSheet("background-color: transparent;")
            year_layout = QHBoxLayout(year_container)
            year_layout.addWidget(year_combo)
            year_layout.setContentsMargins(6, 0, 6, 0)
            year_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setCellWidget(row_position, 0, year_container)

            semester_combo = QComboBox()
            semester_combo.addItems(['1', '2'])
            semester_combo.setCurrentText(str(data['Семестр']))
            semester_combo.setStyleSheet(table_combo_style)
            semester_container = QWidget()
            semester_container.setStyleSheet("background-color: transparent;")
            semester_layout = QHBoxLayout(semester_container)
            semester_layout.addWidget(semester_combo)
            semester_layout.setContentsMargins(6, 0, 6, 0)
            semester_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setCellWidget(row_position, 1, semester_container)

            type_combo = QComboBox()
            type_combo.addItems(['Т', 'Э', 'П', 'У', 'ПА', 'ГИА', 'Г', 'Д', 'К'])
            type_combo.setCurrentText(data['Тип'])
            type_combo.setStyleSheet(table_combo_style)
            type_container = QWidget()
            type_container.setStyleSheet("background-color: transparent;")
            type_layout = QHBoxLayout(type_container)
            type_layout.addWidget(type_combo)
            type_layout.setContentsMargins(6, 0, 6, 0)
            type_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setCellWidget(row_position, 2, type_container)

            weeks_item = QTableWidgetItem(str(data['Недели']))
            weeks_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            font = QFont()
            font.setPointSize(16)
            weeks_item.setFont(font)
            self.table.setItem(row_position, 3, weeks_item)

            self.table.setRowHeight(row_position, 82)

        self.update_weeks_total()

    def get_table_data(self):
        data = []
        for row in range(self.table.rowCount()):
            year_container = self.table.cellWidget(row, 0)
            semester_container = self.table.cellWidget(row, 1)
            type_container = self.table.cellWidget(row, 2)
            weeks_item = self.table.item(row, 3)

            if year_container and semester_container and type_container and weeks_item:
                try:
                    year_combo = year_container.findChild(QComboBox)
                    semester_combo = semester_container.findChild(QComboBox)
                    type_combo = type_container.findChild(QComboBox)

                    if year_combo and semester_combo and type_combo:
                        weeks = float(weeks_item.text())

                        # Валидация количества недель
                        if weeks < 0:
                            QMessageBox.warning(self, 'Ошибка валидации',
                                              f'Ошибка в строке {row + 1}: Количество недель не может быть отрицательным!\n'
                                              f'Введено: {weeks}')
                            return None

                        if weeks > 53:
                            QMessageBox.warning(self, 'Ошибка валидации',
                                              f'Ошибка в строке {row + 1}: Количество недель не может превышать 53!\n'
                                              f'Введено: {weeks}')
                            return None

                        data.append({
                            'Год': int(year_combo.currentText()),
                            'Семестр': int(semester_combo.currentText()),
                            'Тип': type_combo.currentText(),
                            'Недели': weeks
                        })
                except ValueError:
                    QMessageBox.warning(self, 'Ошибка валидации',
                                      f'Ошибка в строке {row + 1}: Некорректное значение в столбце "Недели"!\n'
                                      f'Введите число (можно дробное, например: 2.5)')
                    return None
        return data

    def generate_schedule(self):
        periods_data = self.get_table_data()

        if not periods_data:
            QMessageBox.warning(self, 'Внимание', 'Добавьте периоды обучения')
            return

        try:
            self.generated_schedule = self.app.generate_schedule(periods_data, self.start_year)

            self.preview_table.setRowCount(0)
            for period in self.generated_schedule:
                row_position = self.preview_table.rowCount()
                self.preview_table.insertRow(row_position)

                items = [
                    str(period['year']),
                    str(period['semester']),
                    period['type'],
                    f"{period['weeks']:.1f}",
                    period['start_date'].strftime('%d.%m.%Y'),
                    period['end_date'].strftime('%d.%m.%Y'),
                    str(len(period['days']))
                ]

                for col, text in enumerate(items):
                    item = QTableWidgetItem(text)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                    font = QFont()
                    font.setPointSize(16)
                    item.setFont(font)
                    self.preview_table.setItem(row_position, col, item)

                self.preview_table.setRowHeight(row_position, 60)

            self.preview_section.setVisible(True)
            self.download_btn.setEnabled(True)

            QMessageBox.information(self, 'Успех',
                                    f'✅ График создан!\n\n'
                                    f'📊 Периодов: {len(self.generated_schedule)}\n'
                                    f'📅 Недель: {sum(p["weeks"] for p in self.generated_schedule):.1f}\n'
                                    f'📝 Рабочих дней: {sum(len(p["days"]) for p in self.generated_schedule)}')

        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', f'Ошибка при генерации:\n{str(e)}')

    def download_excel(self):
        if not self.generated_schedule:
            QMessageBox.warning(self, 'Внимание', 'Сначала сгенерируйте график')
            return

        filename, _ = QFileDialog.getSaveFileName(
            self,
            'Сохранить Excel файл',
            f'график_{self.start_year}-{self.start_year + (2 if "Ординатура" in self.program_type else 3)}.xlsx',
            'Excel Files (*.xlsx)'
        )

        if filename:
            try:
                wb = self.app.create_excel_file(self.generated_schedule, self.start_year, self.program_type)
                wb.save(filename)
                QMessageBox.information(self, 'Успех', f'✅ Файл сохранен:\n{filename}')
            except Exception as e:
                QMessageBox.critical(self, 'Ошибка', f'Ошибка при сохранении:\n{str(e)}')


def main():
    app = QApplication(sys.argv)

    font = QFont()
    font.setPointSize(10)
    app.setFont(font)

    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == '__main__':
    main()